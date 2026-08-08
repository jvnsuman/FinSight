"""
Tests for backend/services/report_service.py - the server-side Excel
report generator.

These specifically verify the reason this endpoint exists server-side at
all: openpyxl must both READ the two native charts already in
backend/templates/monthly_report_template.xlsx and WRITE the result back
out with those same two charts still present. (Verified separately,
outside this test suite, that ExcelJS crashes trying to even read that
template, and SheetJS reads it fine but silently drops both charts on
save - this is why report generation happens in Python, not the browser.)
"""
import os
import sys
from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..")))

from backend.services.report_service import generate_monthly_report_excel, MAX_CATEGORIES, MAX_DAYS


def _save_and_reload(wb):
    """Round-trip through bytes, the same way the FastAPI endpoint does via StreamingResponse."""
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


class TestChartsSurviveGeneration:
    def test_report_sheet_has_exactly_two_charts(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        """
        The core thing this whole feature depends on: the two native
        charts baked into the template must still be there after
        openpyxl loads the template, writes this month's data into it,
        and saves. If this regresses, the "chart" in the exported Excel
        file silently becomes just a data sheet again.
        """
        user = make_user()
        account = make_account(user)
        cat = make_category(user, category_type="expense", category_name="Food")
        make_transaction(
            user, account, cat,
            transaction_type="expense", amount=Decimal("500.00"),
            transaction_date=date(2026, 7, 10),
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)

        report_sheet = reloaded["Report"]
        assert len(report_sheet._charts) == 2

    def test_chart_types_are_pie_and_bar(
        self, db_session, make_user
    ):
        user = make_user()
        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)

        chart_types = sorted(type(c).__name__ for c in reloaded["Report"]._charts)
        assert chart_types == ["BarChart", "PieChart"]


class TestCategoryBreakdownData:
    def test_expense_categories_written_to_chart_data_sheet(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        user = make_user()
        account = make_account(user)
        food_cat = make_category(user, category_type="expense", category_name="Food")
        transport_cat = make_category(user, category_type="expense", category_name="Transport")

        make_transaction(
            user, account, food_cat,
            transaction_type="expense", amount=Decimal("300.00"),
            transaction_date=date(2026, 7, 5),
        )
        make_transaction(
            user, account, transport_cat,
            transaction_type="expense", amount=Decimal("150.00"),
            transaction_date=date(2026, 7, 12),
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        data_sheet = reloaded["ChartData"]

        rows = [
            (data_sheet.cell(row=r, column=1).value, data_sheet.cell(row=r, column=2).value)
            for r in range(2, 2 + MAX_CATEGORIES)
            if data_sheet.cell(row=r, column=1).value is not None
        ]
        assert ("Food", 300.0) in rows
        assert ("Transport", 150.0) in rows

    def test_income_transactions_excluded_from_category_breakdown(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        """Only expenses should appear in the pie chart's data - income has
        its own line in the daily bar chart instead."""
        user = make_user()
        account = make_account(user)
        income_cat = make_category(user, category_type="income", category_name="Salary")

        make_transaction(
            user, account, income_cat,
            transaction_type="income", amount=Decimal("50000.00"),
            transaction_date=date(2026, 7, 1),
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        data_sheet = reloaded["ChartData"]

        category_names = [
            data_sheet.cell(row=r, column=1).value
            for r in range(2, 2 + MAX_CATEGORIES)
        ]
        assert "Salary" not in category_names

    def test_more_than_template_capacity_folds_into_other(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        """
        The template has a fixed MAX_CATEGORIES rows. A user with more
        distinct expense categories than that must not crash or silently
        drop data - the smallest categories should fold into one "Other"
        row so the total still reconciles.
        """
        user = make_user()
        account = make_account(user)

        # MAX_CATEGORIES + 3 distinct categories, each with a different amount
        for i in range(MAX_CATEGORIES + 3):
            cat = make_category(user, category_type="expense", category_name=f"Cat{i}")
            make_transaction(
                user, account, cat,
                transaction_type="expense", amount=Decimal(str(10 + i)),
                transaction_date=date(2026, 7, 1),
            )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        data_sheet = reloaded["ChartData"]

        rows = [
            (data_sheet.cell(row=r, column=1).value, data_sheet.cell(row=r, column=2).value)
            for r in range(2, 2 + MAX_CATEGORIES)
            if data_sheet.cell(row=r, column=1).value is not None
        ]
        # Must not exceed the template's row budget
        assert len(rows) <= MAX_CATEGORIES
        # The overflow must be folded, not dropped - so an "Other" bucket appears
        names = [n for n, _ in rows]
        assert "Other" in names


class TestDailyIncomeExpenseData:
    def test_daily_totals_grouped_by_calendar_day(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        user = make_user()
        account = make_account(user)
        income_cat = make_category(user, category_type="income")
        expense_cat = make_category(user, category_type="expense")

        make_transaction(
            user, account, income_cat,
            transaction_type="income", amount=Decimal("1000.00"),
            transaction_date=date(2026, 7, 1),
        )
        make_transaction(
            user, account, expense_cat,
            transaction_type="expense", amount=Decimal("200.00"),
            transaction_date=date(2026, 7, 1),
        )
        make_transaction(
            user, account, expense_cat,
            transaction_type="expense", amount=Decimal("50.00"),
            transaction_date=date(2026, 7, 1),
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        data_sheet = reloaded["ChartData"]

        # Day 1 = the first daily row of the block
        from backend.services.report_service import DAILY_START_ROW
        assert data_sheet.cell(row=DAILY_START_ROW, column=2).value == 1000.0  # income
        assert data_sheet.cell(row=DAILY_START_ROW, column=3).value == 250.0   # 200 + 50 expense


class TestSummaryAndLedger:
    def test_summary_sheet_totals(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        user = make_user()
        account = make_account(user)
        income_cat = make_category(user, category_type="income")
        expense_cat = make_category(user, category_type="expense")

        make_transaction(
            user, account, income_cat,
            transaction_type="income", amount=Decimal("5000.00"),
            transaction_date=date(2026, 7, 3),
        )
        make_transaction(
            user, account, expense_cat,
            transaction_type="expense", amount=Decimal("1200.00"),
            transaction_date=date(2026, 7, 8),
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        summary = reloaded["Summary"]

        values = {
            summary.cell(row=r, column=1).value: summary.cell(row=r, column=2).value
            for r in range(2, 6)
        }
        assert values["Total Income"] == 5000.0
        assert values["Total Expense"] == 1200.0
        assert values["Net Savings"] == 3800.0

    def test_ledger_lists_every_transaction_in_the_month(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        user = make_user()
        account = make_account(user)
        cat = make_category(user, category_type="expense")

        make_transaction(
            user, account, cat,
            transaction_type="expense", amount=Decimal("42.50"),
            transaction_date=date(2026, 7, 15), description="Groceries",
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        ledger = reloaded["Transaction Ledger"]

        assert ledger.cell(row=2, column=2).value == "Groceries"
        assert ledger.cell(row=2, column=5).value == 42.5

    def test_transactions_outside_the_month_are_excluded(
        self, db_session, make_user, make_account, make_category, make_transaction
    ):
        user = make_user()
        account = make_account(user)
        cat = make_category(user, category_type="expense")

        make_transaction(
            user, account, cat,
            transaction_type="expense", amount=Decimal("99.00"),
            transaction_date=date(2026, 6, 30),  # June, not July
        )

        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 1))
        reloaded = _save_and_reload(wb)
        summary = reloaded["Summary"]

        total_expense = summary.cell(row=3, column=2).value
        assert total_expense == 0.0


class TestReportHeader:
    def test_month_label_written_to_report_sheet(self, db_session, make_user):
        user = make_user()
        wb = generate_monthly_report_excel(db_session, user.user_id, date(2026, 7, 15))
        reloaded = _save_and_reload(wb)

        assert reloaded["Report"]["C3"].value == "July 2026"
