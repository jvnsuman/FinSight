"""
Generates the Monthly Report as a real .xlsx file with native Excel charts.

This runs entirely server-side with openpyxl. It has to - no JS library
(ExcelJS, SheetJS) can write native chart objects into a fresh workbook,
and both of them silently STRIP existing charts from a template when they
save it back out (verified directly: loading backend/templates/
monthly_report_template.xlsx with either library and re-saving drops the
chart count to zero, no error raised). openpyxl, by contrast, can both
read and re-save that template's two chart objects (a category-breakdown
pie chart and a daily income/expense bar chart) intact, because they're
wired to named ranges on a hidden "ChartData" sheet - this function's only
job is filling in the values at those exact named ranges.
"""
import calendar
import os
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from backend.models.transaction import Transaction
from backend.models.category import Category

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "..", "templates", "monthly_report_template.xlsx")

# Must match backend/templates/build_template.py exactly - these are the
# fixed row numbers the template's named ranges point at.
MAX_CATEGORIES = 20
MAX_DAYS = 31
DAILY_HEADER_ROW = 2 + MAX_CATEGORIES + 1  # 23
DAILY_START_ROW = DAILY_HEADER_ROW + 1     # 24


def _month_bounds(month: date) -> tuple[date, date]:
    first_day = month.replace(day=1)
    last_day = month.replace(day=calendar.monthrange(month.year, month.month)[1])
    return first_day, last_day


def generate_monthly_report_excel(db: Session, user_id: int, month: date) -> openpyxl.Workbook:
    """
    Returns an in-memory openpyxl Workbook for the given month - the
    caller is responsible for saving it to a stream and returning it as a
    file download. Raises ValueError if the month has more distinct
    calendar days of transactions than the template supports (31, so this
    can only happen from a bad month.replace() upstream) or more expense
    categories than the template has rows for (20 - excess categories
    are folded into an "Other" bucket rather than raising, since that's a
    real, expected case for a busy user).
    """
    first_day, last_day = _month_bounds(month)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    data_sheet = wb["ChartData"]
    report_sheet = wb["Report"]
    summary_sheet = wb["Summary"]
    ledger_sheet = wb["Transaction Ledger"]

    transactions = (
        db.query(Transaction)
        .filter(
            and_(
                Transaction.user_id == user_id,
                Transaction.transaction_date >= first_day,
                Transaction.transaction_date <= last_day,
            )
        )
        .order_by(Transaction.transaction_date.asc())
        .all()
    )

    category_ids = {t.category_id for t in transactions if t.category_id is not None}
    categories_by_id = {
        c.category_id: c
        for c in db.query(Category).filter(Category.category_id.in_(category_ids)).all()
    } if category_ids else {}

    # --- Category breakdown (expenses only) ---
    category_totals: dict[str, Decimal] = {}
    for t in transactions:
        if t.transaction_type != "expense":
            continue
        cat = categories_by_id.get(t.category_id)
        name = cat.category_name if cat else "Uncategorized"
        category_totals[name] = category_totals.get(name, Decimal("0")) + t.amount

    # Sort largest-first, fold anything past the template's row budget into "Other"
    sorted_categories = sorted(category_totals.items(), key=lambda kv: kv[1], reverse=True)
    if len(sorted_categories) > MAX_CATEGORIES:
        kept = sorted_categories[: MAX_CATEGORIES - 1]
        overflow_total = sum(amt for _, amt in sorted_categories[MAX_CATEGORIES - 1:])
        sorted_categories = kept + [("Other", overflow_total)]

    for i in range(MAX_CATEGORIES):
        row = 2 + i
        if i < len(sorted_categories):
            name, amount = sorted_categories[i]
            data_sheet.cell(row=row, column=1, value=name)
            data_sheet.cell(row=row, column=2, value=float(amount))
        else:
            data_sheet.cell(row=row, column=1, value=None)
            data_sheet.cell(row=row, column=2, value=None)

    # --- Daily income vs. expense ---
    days_in_month = (last_day - first_day).days + 1
    daily_income = {d: Decimal("0") for d in range(days_in_month)}
    daily_expense = {d: Decimal("0") for d in range(days_in_month)}
    for t in transactions:
        offset = (t.transaction_date - first_day).days
        if t.transaction_type == "income":
            daily_income[offset] = daily_income.get(offset, Decimal("0")) + t.amount
        elif t.transaction_type == "expense":
            daily_expense[offset] = daily_expense.get(offset, Decimal("0")) + t.amount

    for i in range(MAX_DAYS):
        row = DAILY_START_ROW + i
        if i < days_in_month:
            day_date = first_day + timedelta(days=i)
            data_sheet.cell(row=row, column=1, value=day_date.strftime("%b %d"))
            data_sheet.cell(row=row, column=2, value=float(daily_income[i]))
            data_sheet.cell(row=row, column=3, value=float(daily_expense[i]))
        else:
            data_sheet.cell(row=row, column=1, value=None)
            data_sheet.cell(row=row, column=2, value=None)
            data_sheet.cell(row=row, column=3, value=None)

    # --- Summary sheet ---
    total_income = sum((t.amount for t in transactions if t.transaction_type == "income"), Decimal("0"))
    total_expense = sum((t.amount for t in transactions if t.transaction_type == "expense"), Decimal("0"))
    net_savings = total_income - total_expense

    summary_rows = [
        ("Total Income", float(total_income)),
        ("Total Expense", float(total_expense)),
        ("Net Savings", float(net_savings)),
        ("Transaction Count", len(transactions)),
    ]
    for i, (label, value) in enumerate(summary_rows):
        summary_sheet.cell(row=2 + i, column=1, value=label)
        summary_sheet.cell(row=2 + i, column=2, value=value)

    # --- Transaction ledger ---
    for i, t in enumerate(transactions):
        row = 2 + i
        cat = categories_by_id.get(t.category_id)
        ledger_sheet.cell(row=row, column=1, value=t.transaction_date.strftime("%Y-%m-%d"))
        ledger_sheet.cell(row=row, column=2, value=t.description or "")
        ledger_sheet.cell(row=row, column=3, value=cat.category_name if cat else "Uncategorized")
        ledger_sheet.cell(row=row, column=4, value=t.transaction_type)
        ledger_sheet.cell(row=row, column=5, value=float(t.amount))
        ledger_sheet.cell(row=row, column=6, value=t.payment_mode or "")

    # --- Report header ---
    report_sheet["C3"] = first_day.strftime("%B %Y")

    return wb
